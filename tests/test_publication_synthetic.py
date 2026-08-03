"""Tests d'acceptation hors ligne pour la ponderation et la publication."""

from __future__ import annotations

import importlib
import io
from dataclasses import replace
from pathlib import Path

import numpy as np
import polars as pl
from openpyxl import load_workbook

from cnps.config import StatDef, load_config

ROOT = Path(__file__).resolve().parents[1]
CFG = load_config(ROOT / "config/settings.yaml", ROOT / "config/dimensions.yaml")
weighting = importlib.import_module("cnps.09_ponderation_finale")
estimation = importlib.import_module("cnps.10_estimation_indicateurs")
validation = importlib.import_module("cnps.11_validation_qualite")
export = importlib.import_module("cnps.12_export_excel")
pipeline = importlib.import_module("cnps.pipeline")


def _publication_cfg():
    return replace(
        CFG,
        estimation=replace(
            CFG.estimation,
            min_distinct_individuals=1,
            min_distinct_employers=1,
            max_employer_wage_share=1.0,
        ),
    )


def test_two_stage_weight_is_response_over_product_of_propensities() -> None:
    df = pl.DataFrame(
        {
            "D_JT": [1, 1, 0, 1],
            "S_IJT": [1, 0, 0, 1],
            "P_HAT_JT": [0.5] * 4,
            "Q_HAT_IJT": [0.8, 0.8, 1.0, 0.8],
            "W_JT": [2.0] * 4,
            "W_INDIV": [1.25, 1.25, 1.0, 1.25],
            "DANS_UNIVERS_RISQUE": [1, 1, 1, 0],
        }
    )
    weighting._validate_input_weights(
        df.filter(pl.col("DANS_UNIVERS_RISQUE") == 1),
        CFG,
    )
    raw, final = weighting._compute_two_stage_weights(df, CFG)
    np.testing.assert_allclose(raw, [2.5, 0.0, 0.0, 0.0])
    np.testing.assert_allclose(final, raw)


def test_disclosure_counts_distinct_positive_weight_contributors() -> None:
    repeated = pl.DataFrame(
        {
            "ID_INDIV": ["A", "B"] * 15,
            "ID_EMPLOYEUR": ["E1", "E2"] * 15,
            "SALAIRE_BRUT_ESTIME_AU_MOIS": [100_000.0] * 30,
            "W_FINAL": [1.0] * 30,
        }
    )
    status, _ = estimation._disclosure_status(repeated, CFG, "W_FINAL")
    assert status.startswith("primaire:")
    assert "individus" in status
    assert "employeurs" in status

    zero_weight = repeated.with_columns(pl.lit(0.0).alias("W_FINAL"))
    status_zero, _ = estimation._disclosure_status(
        zero_weight,
        _publication_cfg(),
        "W_FINAL",
    )
    assert status_zero.startswith("primaire:")


def test_disclosure_dominance_uses_unweighted_observed_wage_mass() -> None:
    df = pl.DataFrame(
        {
            "ID_INDIV": [f"I{i}" for i in range(30)],
            "ID_EMPLOYEUR": ["BIG"] * 28 + ["E2", "E3"],
            "SALAIRE_BRUT_ESTIME_AU_MOIS": [1_000_000.0] * 28 + [100_000.0, 100_000.0],
            "W_FINAL": [1.0] * 30,
        }
    )
    status, _ = estimation._disclosure_status(df, CFG, "W_FINAL")
    assert "dominance" in status


def test_each_statistic_uses_its_configured_variable() -> None:
    df = pl.DataFrame(
        {
            "ID_INDIV": ["A", "B"],
            "ID_EMPLOYEUR": ["E1", "E2"],
            "SALAIRE_BRUT_ESTIME_AU_MOIS": [100.0, 1_000.0],
            "SALAIRE_BRUT_ESTIME_AU_MOIS_W": [100.0, 200.0],
            "W_FINAL": [1.0, 1.0],
        }
    )
    statistics = [
        StatDef(
            name="mean",
            label="Moyenne",
            function="weighted_mean",
            variable="SALAIRE_BRUT_ESTIME_AU_MOIS_W",
        ),
        StatDef(
            name="max",
            label="Maximum observe",
            function="max",
            variable="SALAIRE_BRUT_ESTIME_AU_MOIS",
        ),
    ]
    result = estimation._estimate_group(
        df,
        "W_FINAL",
        statistics,
        _publication_cfg(),
    )
    assert result["mean"] == 150.0
    assert result["max"] == 1_000.0
    assert result["inference_status"] == "POINT_ONLY_F1_PENDING"


def test_secondary_suppression_masks_the_smallest_remaining_cell() -> None:
    rows = [
        {
            "group": "A",
            "mean": None,
            "suppression_status": "primaire:individus",
            "_wage_mass_observed": 10.0,
        },
        {
            "group": "B",
            "mean": 200.0,
            "suppression_status": "publiee",
            "_wage_mass_observed": 20.0,
        },
        {
            "group": "C",
            "mean": 300.0,
            "suppression_status": "publiee",
            "_wage_mass_observed": 30.0,
        },
    ]
    stat = StatDef("mean", "Moyenne", "weighted_mean", "salary")
    estimation._apply_secondary_suppression(rows, [stat])
    assert rows[1]["mean"] is None
    assert rows[1]["suppression_status"] == "secondaire"
    assert rows[2]["mean"] == 300.0


def test_secondary_suppression_is_applied_separately_by_month() -> None:
    rows = [
        {
            "group": "2024-01 / A",
            "mean": None,
            "suppression_status": "primaire:individus",
            "_wage_mass_observed": 10.0,
            "_secondary_partition": "2024-01",
        },
        {
            "group": "2024-01 / B",
            "mean": 20.0,
            "suppression_status": "publiee",
            "_wage_mass_observed": 20.0,
            "_secondary_partition": "2024-01",
        },
        {
            "group": "2024-02 / A",
            "mean": None,
            "suppression_status": "primaire:employeurs",
            "_wage_mass_observed": 30.0,
            "_secondary_partition": "2024-02",
        },
        {
            "group": "2024-02 / B",
            "mean": 40.0,
            "suppression_status": "publiee",
            "_wage_mass_observed": 40.0,
            "_secondary_partition": "2024-02",
        },
    ]
    stat = StatDef("mean", "Moyenne", "weighted_mean", "salary")
    estimation._apply_secondary_suppression(rows, [stat])
    assert rows[1]["suppression_status"] == "secondaire"
    assert rows[3]["suppression_status"] == "secondaire"


def test_validation_rejects_intervals_while_f1_is_pending() -> None:
    base = pl.DataFrame(
        {
            "dimension": ["National"],
            "group": ["Total"],
            "n_obs": [100.0],
            "mean": [100_000.0],
            "variance": [1_000.0],
            "min": [75_000.0],
            "p10": [80_000.0],
            "q1": [90_000.0],
            "median": [100_000.0],
            "q3": [110_000.0],
            "p90": [120_000.0],
            "max": [130_000.0],
            "gini": [0.20],
            "suppression_status": ["publiee"],
            "inference_status": ["POINT_ONLY_F1_PENDING"],
        }
    )
    assert validation.valider_estimation(CFG, base).is_valid
    with_interval = base.with_columns(pl.lit(90_000.0).alias("mean_ci_lower"))
    report = validation.valider_estimation(CFG, with_interval)
    assert not report.is_valid
    assert any(issue.check == "intervals_forbidden" for issue in report.errors)


def test_validation_rejects_nonfinite_or_unmasked_publication_cells() -> None:
    values = {
        "dimension": ["National", "Secteur"],
        "group": ["Total", "A"],
        "n_obs": [100.0, None],
        "mean": [float("nan"), 100_000.0],
        "variance": [1_000.0, None],
        "min": [75_000.0, None],
        "p10": [80_000.0, None],
        "q1": [90_000.0, None],
        "median": [100_000.0, None],
        "q3": [110_000.0, None],
        "p90": [120_000.0, None],
        "max": [130_000.0, None],
        "gini": [0.20, None],
        "suppression_status": ["publiee", "primaire:individus"],
        "inference_status": ["POINT_ONLY_F1_PENDING"] * 2,
    }
    report = validation.valider_estimation(CFG, pl.DataFrame(values))
    assert any(issue.check == "nonfinite_published_statistics" for issue in report.errors)

    leaked = pl.DataFrame(values).with_columns(
        pl.when(pl.col("group") == "A").then(100_000.0).otherwise(pl.col("mean")).alias("mean")
    )
    report = validation.valider_estimation(CFG, leaked)
    assert any(issue.check == "unsafely_unmasked_statistics" for issue in report.errors)


def test_saved_response_diagnostics_are_revalidated() -> None:
    model_data = {
        "schema_version": 1,
        "diagnostics_oof": {
            "auc": 0.55,
            "calibration_in_large": 0.01,
            "brier": 0.20,
            "max_abs_smd": 0.10,
            "propensity_min": 0.20,
            "propensity_max": 0.80,
            "n_splits": 5,
        },
    }
    report = validation.ValidationReport()
    validation._ajouter_diagnostics_modele_reponse(
        report,
        model_data,
        CFG,
        label="Modele test",
        code_suffix="_test",
    )
    assert report.is_valid
    assert report.warnings

    model_data["diagnostics_oof"]["calibration_in_large"] = 1.0
    report = validation.ValidationReport()
    validation._ajouter_diagnostics_modele_reponse(
        report,
        model_data,
        CFG,
        label="Modele test",
        code_suffix="_test",
    )
    assert not report.is_valid
    assert any(issue.check == "failed_diagnostics_test" for issue in report.errors)


def test_excel_masks_nonfinite_values_and_formats_gini(monkeypatch) -> None:
    results = pl.DataFrame(
        {
            "dimension": ["National"],
            "group": ["Total"],
            "gini": [0.3714],
            "mean": [float("nan")],
            "suppression_status": ["publiee"],
            "inference_status": ["POINT_ONLY_F1_PENDING"],
        }
    )
    payload: dict[str, bytes] = {}

    def capture(_cfg, _bucket, _object, write_fn):
        buffer = io.BytesIO()
        write_fn(buffer)
        payload["xlsx"] = buffer.getvalue()

    monkeypatch.setattr(export, "write_workbook", capture)
    export.exporter_indicateurs(CFG, results, validation.ValidationReport())

    workbook = load_workbook(io.BytesIO(payload["xlsx"]), data_only=True)
    sheet = workbook["National"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.cell(2, headers["Coefficient de Gini"]).value == 0.3714
    assert sheet.cell(2, headers["Coefficient de Gini"]).number_format == "0.000"
    assert sheet.cell(2, headers["Salaire moyen"]).value == "—"


def test_publication_dag_excludes_experimental_imputation() -> None:
    assert all("IMPUTATION" not in stage.name for stage in pipeline.Stage)
